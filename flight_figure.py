# 20170709
# Flora Tsai

import sys
import graphviz as gv
import openpyxl
import argparse

max_column = 10
col_index = {'Flight #': 0, 'Departure': 0, 'Arrival': 0, 'Duration': 0}

styles = {
	'graph':{
		'fontsize': '16',
		'rankdir': 'LR'
	},
	'cluster': {
		'fontsize': '12'
	},
	'node':{
		'shape': 'box',
		'style': 'filled'
	},
	'edge':{
		'fontsize': '12'
	}
}

def generate_datasets(input_filename, datasets):

	document = openpyxl.load_workbook(input_filename)
	sheet = document.get_sheet_by_name('Sheet1')
	# sheet = document.__getitem__('Sheet1')

	# Position target columns for data processing
	for i in range(1, max_column):
		if sheet.cell(row=1, column=i).value == 'Flight number':
			col_index['Flight #'] = i
		if sheet.cell(row=1, column=i).value == 'Departure':
			col_index['Departure'] = i
		if sheet.cell(row=1, column=i).value == 'Arrival':
			col_index['Arrival'] = i
		if sheet.cell(row=1, column=i).value == 'Duration':
			col_index['Duration'] = i

	cur_row = 2
	flight = []
	departure = []
	arrival = []
	duration = []
	
	# Fill data
	while True:
		content = sheet.cell(row=cur_row, column=col_index['Flight #']).value
		if content is None:
			break

		flight.append(content)
		departure.append(sheet.cell(row=cur_row, column=col_index['Departure']).value)
		arrival.append(sheet.cell(row=cur_row, column=col_index['Arrival']).value)
		duration.append(sheet.cell(row=cur_row, column=col_index['Duration']).value)

		cur_row += 1

	datasets['tag'] = flight
	datasets['Departure'] = departure
	datasets['Arrival'] = arrival
	datasets['Duration'] = duration

def add_cluster(graph, cluster_name, dataset, cluster_label, tagset, node_tag):
	with graph.subgraph(name=cluster_name, node_attr=styles['node']) as cluster:
		cluster.attr(label=cluster_label)
		cluster.attr(fontsize=styles['cluster']['fontsize'])
		for i in range(0, len(dataset)):
			cluster.node(tagset[i] + '-' + node_tag, dataset[i])

def make_edges(graph, datasets, depart, arrive):
	for i in range(0, len(datasets['tag'])):
		tag1 = datasets['tag'][i] + '-' + depart
		tag2 = datasets['tag'][i] + '-' + arrive
		graph.edge(tag1, tag2, label=datasets['tag'][i] + ':\n' + datasets['Duration'][i])

def generate_figure(input_filename, filename, figure_label):

	g = gv.Digraph(name=figure_label, filename=filename+'.gv')
	g.attr(label=figure_label)
	g.attr(rankdir = styles['graph']['rankdir'])

	my_datasets = {}
	generate_datasets(input_filename, my_datasets)


	# To draw clusters, make sure cluster_name starts with 'cluster_'
	add_cluster(graph=g, cluster_name='cluster_depart', dataset=my_datasets['Departure'],
		        cluster_label='Departure', tagset=my_datasets['tag'], node_tag='depart')
	
	add_cluster(graph=g, cluster_name='cluster_arrive', dataset=my_datasets['Arrival'],
		        cluster_label='Arrival', tagset=my_datasets['tag'], node_tag='arrive')

	make_edges(graph=g, datasets=my_datasets, depart='depart', arrive='arrive')

	# View graph
	g.view()
	# Print dot format in terminal
	print(g.source)


if __name__ == '__main__':
	if len(sys.argv) == 4:
		generate_figure(sys.argv[1], sys.argv[2], sys.argv[3])
	else:
		print('Usage: python flight_figure <input_filename> <output_filename> <figure_label>')

###EOF		