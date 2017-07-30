# 20170729
# Flora Tsai

import graphviz as gv
import openpyxl
import argparse

max_column = 10
col_index = {'Flight #': 0, 'Departure': 0, 'Arrival': 0, 'Duration': 0}

styles = {
	'graph':{
		'fontsize': '16',
		'rankdir': 'LR'
	},
	'cluster': {
		'fontsize': '12'
	},
	'node':{
		'shape': 'box',
		'style': 'filled'
	},
	'edge':{
		'fontsize': '12'
	}
}

def generate_datasets(input_filename, datasets, _departure, _arrival, search):

	document = openpyxl.load_workbook(input_filename)
	sheet = document.get_sheet_by_name('Sheet1')
	# sheet = document.__getitem__('Sheet1')

	# Position target columns for data processing
	for i in range(1, max_column):
		if sheet.cell(row=1, column=i).value == 'Flight number':
			col_index['Flight #'] = i
		if sheet.cell(row=1, column=i).value == 'Departure':
			col_index['Departure'] = i
		if sheet.cell(row=1, column=i).value == 'Arrival':
			col_index['Arrival'] = i
		if sheet.cell(row=1, column=i).value == 'Duration':
			col_index['Duration'] = i

	cur_row = 2
	flight = []
	departure = []
	arrival = []
	duration = []
	
	# Fill data
	while True:
		content = sheet.cell(row=cur_row, column=col_index['Flight #']).value
		if content is None:
			break

		dep = sheet.cell(row=cur_row, column=col_index['Departure']).value
		arr = sheet.cell(row=cur_row, column=col_index['Arrival']).value

		if search != '' and dep != search and arr != search:
			cur_row += 1
			continue
			
		if (_departure != '' and dep != _departure) or (_arrival != '' and arr != _arrival):
			cur_row += 1
			continue

		flight.append(content)
		departure.append(dep)
		arrival.append(arr)
		duration.append(sheet.cell(row=cur_row, column=col_index['Duration']).value)

		cur_row += 1

	datasets['tag'] = flight
	datasets['Departure'] = departure
	datasets['Arrival'] = arrival
	datasets['Duration'] = duration

def add_cluster(graph, cluster_name, dataset, cluster_label, tagset, node_tag):
	with graph.subgraph(name=cluster_name, node_attr=styles['node']) as cluster:
		cluster.attr(label=cluster_label)
		cluster.attr(fontsize=styles['cluster']['fontsize'])
		for i in range(0, len(dataset)):
			cluster.node(tagset[i] + '-' + node_tag, dataset[i])

def make_edges(graph, datasets, depart, arrive):
	for i in range(0, len(datasets['tag'])):
		tag1 = datasets['tag'][i] + '-' + depart
		tag2 = datasets['tag'][i] + '-' + arrive
		graph.edge(tag1, tag2, label=datasets['tag'][i] + ':\n' + datasets['Duration'][i])

def generate_figure(input_filename, output_filename, figure_label, departure, arrival, search):

	g = gv.Digraph(name=figure_label, filename=output_filename+'.gv')
	g.attr(label=figure_label)
	g.attr(rankdir = styles['graph']['rankdir'])

	my_datasets = {}
	generate_datasets(input_filename, my_datasets, departure, arrival, search)


	# To draw clusters, make sure cluster_name starts with 'cluster_'
	add_cluster(graph=g, cluster_name='cluster_depart', dataset=my_datasets['Departure'],
		        cluster_label='Departure', tagset=my_datasets['tag'], node_tag='depart')
	
	add_cluster(graph=g, cluster_name='cluster_arrive', dataset=my_datasets['Arrival'],
		        cluster_label='Arrival', tagset=my_datasets['tag'], node_tag='arrive')

	make_edges(graph=g, datasets=my_datasets, depart='depart', arrive='arrive')

	# View graph
	g.view()
	# Print dot format in terminal
	print(g.source)


if __name__ == '__main__':
	departure = ''
	arrival = ''
	search = ''

	# Arg parser
	parser = argparse.ArgumentParser()
	parser.add_argument('input_filename')
	parser.add_argument('output_filename')
	parser.add_argument('figure_label')
	parser.add_argument('-d', '--departure', help='Departure')
	parser.add_argument('-a', '--arrival', help='Arrival')
	parser.add_argument('-s', '--search', help='Departure or Arrival')
	args = parser.parse_args()
	if args.departure:
		departure = args.departure
	if args.arrival:
		arrival = args.arrival
	if args.search:
		search = args.search
		departure = ''
		arrival = ''
	
	generate_figure(args.input_filename, args.output_filename, args.figure_label, departure, arrival, search)

###EOF		